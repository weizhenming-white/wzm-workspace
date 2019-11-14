#!/usr/bin/env python
# -*- coding:utf-8 -*-
"""
Author:魏振明
time:2019-05-15 17:39
brief:简单处理测试记录文档
"""
import os
import time
from openpyxl import Workbook


path = "/opt/holo-sim/bin/"
filename = "timer-2019-05-15.txt"

# 获取当天日期
now = int(time.time())
timeArray = time.localtime(now)
otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
date = otherStyleTime.split(" ")[0]

# 创建一个excel工作簿，注意该工作簿是在内存中创建
wb = Workbook()
# 创建一个新的工作表
ws1 = wb.create_sheet(date)
# ws1.title = 'new title'  #改变工作表名

# #单个单元格操作
# #第一种，写入数据的方法，即直接使用单元格编号写入
# c = ws['A4']
# ws['A1'] = 1
# c = 2
# #第二种，写入数据的方法，Worksheet.cell()方法
# d = ws.cell(row=4, column=2, value=10)


def count_timer():
    num = 1
    with open(path + filename) as f:
        for i in f:
            ws1.cell(row=num, column=1, value=str(i))
            num += 1


if __name__ == '__main__':
    # 保存文件
    # 此操作会覆盖目录中文件，而不提示警告
    wb.save('test.xlsx')

    # count_timer()